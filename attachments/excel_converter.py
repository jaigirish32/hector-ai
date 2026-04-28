"""
attachments/excel_converter.py

Pure in-memory conversion of .xlsx workbooks to a single concatenated CSV
for upload to provider Files APIs. No filesystem writes, no registry coupling.
"""

from __future__ import annotations

import csv
import io
import warnings
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class SheetSummary:
    name: str
    rows_written: int
    cols_written: int
    merged_range_count: int
    empty_formula_cells: int


@dataclass
class ConversionResult:
    csv_bytes: bytes
    sheet_count: int
    sheets: list[SheetSummary] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def total_rows(self) -> int:
        return sum(s.rows_written for s in self.sheets)


def convert_xlsx_to_csv_bytes(xlsx_path: Path) -> ConversionResult:
    """
    Read an .xlsx file and return a single concatenated CSV as UTF-8 bytes.

    Sheets are separated by a '=== Sheet: <name> ===' marker row and a blank
    line. Cached formula values are read (data_only=True); formula cells with
    no cached value appear as empty and are counted in warnings.

    Merged cells flatten to their top-left value (openpyxl default); the rest
    of the merged range becomes blank. Counts are reported in warnings.

    Raises FileNotFoundError if the path does not exist; lets openpyxl raise
    on malformed / non-xlsx files.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.is_file():
        raise FileNotFoundError(xlsx_path)

    # Two loads: data_only=True gives cached values but cannot distinguish a
    # real blank from a formula with no cached value. data_only=False gives us
    # the formula string so we can detect that case for the warning count.
    seen_warnings: dict[str, int] = {}
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        wb_values = load_workbook(filename=str(xlsx_path), data_only=True)
        for w in caught:
            msg = str(w.message)
            seen_warnings[msg] = seen_warnings.get(msg, 0) + 1
    wb_formulas = load_workbook(filename=str(xlsx_path), data_only=False)

    captured: list[str] = []
    for msg, count in seen_warnings.items():
        captured.append(f"openpyxl: {msg}" + (f" (×{count})" if count > 1 else ""))

    buf = io.StringIO(newline="")
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL, lineterminator="\n")

    sheets: list[SheetSummary] = []
    for idx, name in enumerate(wb_values.sheetnames):
        ws_v: Worksheet = wb_values[name]
        ws_f: Worksheet = wb_formulas[name]

        if idx > 0:
            buf.write("\n")
        writer.writerow([f"=== Sheet: {name} ==="])

        rows, cols, empties = _write_sheet(ws_v, ws_f, writer)
        sheets.append(
            SheetSummary(
                name=name,
                rows_written=rows,
                cols_written=cols,
                merged_range_count=len(ws_v.merged_cells.ranges),
                empty_formula_cells=empties,
            )
        )

    csv_bytes = buf.getvalue().encode("utf-8")

    warning_msgs: list[str] = list(captured)
    total_merged = sum(s.merged_range_count for s in sheets)
    if total_merged:
        affected = sum(1 for s in sheets if s.merged_range_count)
        warning_msgs.append(
            f"{total_merged} merged cell range(s) across {affected} sheet(s) "
            f"flattened to top-left value; other merged positions are blank"
        )
    total_empty_f = sum(s.empty_formula_cells for s in sheets)
    if total_empty_f:
        warning_msgs.append(
            f"{total_empty_f} formula cell(s) had no cached value — workbook "
            f"was likely never opened in Excel/LibreOffice since formulas were "
            f"written. Most are typically blank-by-design; if expected data is "
            f"missing in the output, open the original in Excel, save, and re-add."
        )

    return ConversionResult(
        csv_bytes=csv_bytes,
        sheet_count=len(sheets),
        sheets=sheets,
        warnings=warning_msgs,
    )


def _write_sheet(ws_v: Worksheet, ws_f: Worksheet, writer) -> tuple[int, int, int]:
    """Write one sheet's rows. Returns (row_count, col_count, empty_formula_count)."""
    data_rows = list(ws_v.iter_rows(values_only=True))
    formula_rows = list(ws_f.iter_rows(values_only=True))

    # Trim trailing all-blank rows (openpyxl can pad based on formatted-but-empty cells).
    while data_rows and all(_is_blank(c) for c in data_rows[-1]):
        data_rows.pop()
        if formula_rows:
            formula_rows.pop()

    if not data_rows:
        return (0, 0, 0)

    # Last column with any non-blank value across the sheet.
    last_col = 0
    for r in data_rows:
        for i in range(len(r) - 1, -1, -1):
            if not _is_blank(r[i]):
                last_col = max(last_col, i + 1)
                break
    if last_col == 0:
        return (0, 0, 0)

    empties = 0
    for di, drow in enumerate(data_rows):
        frow = formula_rows[di] if di < len(formula_rows) else ()
        out: list[str] = []
        for ci in range(last_col):
            dval = drow[ci] if ci < len(drow) else None
            fval = frow[ci] if ci < len(frow) else None
            if dval is None and isinstance(fval, str) and fval.startswith("="):
                empties += 1
                out.append("")
            else:
                out.append("" if dval is None else str(dval))
        writer.writerow(out)

    return (len(data_rows), last_col, empties)


def _is_blank(v) -> bool:
    return v is None or (isinstance(v, str) and v == "")